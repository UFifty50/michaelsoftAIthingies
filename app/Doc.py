import openpyxl.workbook
import openpyxl.xml
import openpyxl
import xlrd
import csv
import io
openpyxl.xml.DEFUSEDXML = True


def openWorkbook(content: bytes) -> openpyxl.Workbook | xlrd.Book | str:
    try:
        return openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    except Exception:
        try:
            return xlrd.open_workbook(file_contents=content)
        except xlrd.XLRDError:
            return content.decode("utf-8")
    
def xl2csv(content: openpyxl.Workbook | xlrd.Book | str) -> str:
    if isinstance(content, openpyxl.Workbook):
        sheet = content.active
        csvStr = io.StringIO()
        csvWriter = csv.writer(csvStr)
        for row in sheet.iter_rows():
            csvWriter.writerow([cell.value for cell in row])
        return csvStr.getvalue()
    elif isinstance(content, xlrd.Book):
        return "\n".join([",".join([str(cell.value) for cell in row]) for row in content.sheet_by_index(0).get_rows()])
    else:
        return content
    
    
class CsvDoc:
    title: str
    content: str
    
    def __init__(self, title: str, content: io.BufferedReader):
        self.title = title
        
        contentBytes: bytes = content.read()
        contents = openWorkbook(contentBytes)
        self.content = xl2csv(contents)
        
        